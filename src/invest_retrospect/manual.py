"""수동 주식 원장 → 매매일지.

증권사 연동 없이 사용자가 직접 기록한 **매수/매도 이벤트(변화)** 만으로 임의 날짜의
보유 현황·실현손익·체결 내역을 재생(replay)해 `DailyJournalData` 를 만든다.
해외주식 대응을 위해 종목마다 시장/통화를 가지며, 평가용 현재가는 Yahoo Finance 로
자동 조회하되 실패하면 수동 입력값(없으면 원가)으로 폴백한다.
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from invest_retrospect import prices
from invest_retrospect.analyzer import DailyJournalData, Holding, StockPL, Trade
from invest_retrospect.brokers import Broker, broker_info, make_client
from invest_retrospect.config import Config
from invest_retrospect.core import (
    JournalResult,
    _maybe_ai,
    _write_outputs,
    today_ymd,
    ymd_dashed,
)

LogFn = Callable[[str], None]


@dataclass
class LedgerEntry:
    date: str            # YYYYMMDD
    stk_cd: str
    stk_nm: str
    side: str            # '매수' | '매도'
    qty: int
    price: float         # 거래 단가 (해외는 소수 가능)
    market: str = prices.DEFAULT_MARKET   # KOSPI/KOSDAQ/NASDAQ/...
    tag: str = ""        # 사용자 분류 태그 (예: 단타/장투/테마)

    @property
    def currency(self) -> str:
        return prices.currency_of(self.market)

    def to_dict(self) -> dict[str, Any]:
        return {
            "date": self.date, "stk_cd": self.stk_cd, "stk_nm": self.stk_nm,
            "side": self.side, "qty": self.qty, "price": self.price,
            "market": self.market, "tag": self.tag,
        }

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "LedgerEntry":
        return cls(
            date=str(d.get("date", "")).strip(),
            stk_cd=str(d.get("stk_cd", "")).strip(),
            stk_nm=str(d.get("stk_nm", "")).strip(),
            side="매도" if str(d.get("side", "")).strip() == "매도" else "매수",
            qty=int(d.get("qty") or 0),
            price=float(d.get("price") or 0),
            market=str(d.get("market") or prices.DEFAULT_MARKET).strip() or prices.DEFAULT_MARKET,
            tag=str(d.get("tag") or "").strip(),
        )


@dataclass
class Ledger:
    entries: list[LedgerEntry] = field(default_factory=list)
    prices: dict[str, float] = field(default_factory=dict)   # {stk_cd: 수동 현재가}

    def to_dict(self) -> dict[str, Any]:
        return {"entries": [e.to_dict() for e in self.entries], "prices": self.prices}

    @classmethod
    def from_dict(cls, data: Any) -> "Ledger":
        """{entries:[...], prices:{...}} dict 를 Ledger 로. 형식이 다르면 ValueError."""
        if not isinstance(data, dict):
            raise ValueError("올바른 원장 백업 파일이 아닙니다.")
        entries = [
            LedgerEntry.from_dict(d)
            for d in (data.get("entries") or [])
            if isinstance(d, dict)
        ]
        raw_prices = data.get("prices") or {}
        prices_map = {str(k): float(v) for k, v in raw_prices.items() if _is_num(v)}
        return cls(entries=entries, prices=prices_map)


# 다계좌 미사용 시(구버전 파일·신규 사용자)의 기본 계좌 이름
DEFAULT_ACCOUNT = "기본"


@dataclass
class LedgerBook:
    """여러 계좌(이름→원장)를 담는 묶음. dict 삽입순서 = 탭 순서."""

    accounts: dict[str, Ledger] = field(
        default_factory=lambda: {DEFAULT_ACCOUNT: Ledger()}
    )
    active: str = DEFAULT_ACCOUNT   # 현재 선택된 계좌 이름

    def to_dict(self) -> dict[str, Any]:
        return {
            "accounts": {n: l.to_dict() for n, l in self.accounts.items()},
            "active": self.active,
        }

    @classmethod
    def from_dict(cls, data: Any) -> "LedgerBook":
        """신 포맷({accounts,active})·구 포맷({entries,prices}) 모두 수용.

        구 포맷(단일 원장)은 '기본' 계좌 하나로 자동 마이그레이션한다.
        형식이 dict 가 아니면 ValueError.
        """
        if not isinstance(data, dict):
            raise ValueError("올바른 원장 백업 파일이 아닙니다.")
        if isinstance(data.get("accounts"), dict):
            accounts: dict[str, Ledger] = {}
            for name, d in data["accounts"].items():
                if isinstance(d, dict):
                    accounts[str(name)] = Ledger.from_dict(d)
            if not accounts:
                accounts = {DEFAULT_ACCOUNT: Ledger()}
            active = str(data.get("active") or "")
            if active not in accounts:
                active = next(iter(accounts))
            return cls(accounts=accounts, active=active)
        # 구 포맷: 단일 원장 → 기본 계좌로 이관
        return cls(accounts={DEFAULT_ACCOUNT: Ledger.from_dict(data)},
                   active=DEFAULT_ACCOUNT)


# ── 저장/로드 ────────────────────────────────────────────────────────────────
def _ledger_path() -> Path:
    # 지연 import 로 settings_store ↔ manual 순환 의존 회피.
    from invest_retrospect.settings_store import MANUAL_LEDGER_PATH
    return MANUAL_LEDGER_PATH


def load_book() -> LedgerBook:
    """전체 계좌 묶음을 로드. 파일이 없거나 깨지면 빈 묶음(기본 계좌 1개)."""
    path = _ledger_path()
    if not path.is_file():
        return LedgerBook()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return LedgerBook()
    try:
        return LedgerBook.from_dict(data)
    except ValueError:
        return LedgerBook()


def save_book(book: LedgerBook) -> Path:
    """전체 계좌 묶음을 원자적으로 저장한다.

    임시 파일에 먼저 쓴 뒤 os.replace 로 교체해, 저장 도중 충돌/전원차단이
    일어나도 기존 파일이 부분 기록으로 손상되지 않게 한다.
    """
    path = _ledger_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    payload = json.dumps(book.to_dict(), ensure_ascii=False, indent=2)
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(payload)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)   # 같은 디렉토리 → 원자적 교체
    return path


def load_ledger() -> Ledger:
    """활성 계좌의 원장 (CLI·단일 계좌 호환용)."""
    book = load_book()
    return book.accounts.get(book.active) or Ledger()


def save_ledger(ledger: Ledger) -> Path:
    """활성 계좌의 원장만 갱신해 묶음 전체를 저장 (CLI·단일 계좌 호환용)."""
    book = load_book()
    book.accounts[book.active] = ledger
    return save_book(book)


def export_book(book: LedgerBook, dest: Path) -> Path:
    """전체 계좌 묶음을 지정 경로에 JSON 으로 백업한다."""
    dest = Path(dest)
    if dest.parent and not dest.parent.exists():
        dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(
        json.dumps(book.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return dest


def import_book(src: Path) -> LedgerBook:
    """백업 JSON 파일에서 전체 계좌 묶음을 복원한다.

    구버전 단일 원장 백업도 '기본' 계좌로 수용한다. 파일이 없으면 OSError,
    JSON 형식이 아니면 json.JSONDecodeError/ValueError 를 올린다.
    """
    data = json.loads(Path(src).read_text(encoding="utf-8"))
    return LedgerBook.from_dict(data)


def _is_num(v: Any) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


BULK_HEADER = ["거래일", "시장", "종목명", "코드", "구분", "수량", "단가", "태그"]
_BULK_COLS = ", ".join(BULK_HEADER[:7]) + " (+선택: 태그)"


def _valid_ymd(s: str) -> bool:
    s = (s or "").strip()
    if len(s) != 8 or not s.isdigit():
        return False
    from datetime import datetime
    try:
        datetime.strptime(s, "%Y%m%d")
        return True
    except ValueError:
        return False


def _cell_date(v: Any) -> str:
    """엑셀 날짜셀(datetime) 또는 문자열을 YYYYMMDD 로 정규화."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):          # datetime/date
        return v.strftime("%Y%m%d")
    s = str(v).strip()
    return s.replace("-", "").replace("/", "").replace(".", "")


def _cell_code(v: Any, market: str) -> str:
    """코드셀 정규화. 엑셀이 숫자로 읽어 앞자리 0 이 빠진 국내코드는 6자리로 복원."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.isdigit() and prices.currency_of(market) == "KRW" and len(s) < 6:
        s = s.zfill(6)
    return s


def _entry_from_cells(cells: list[Any], rownum: int) -> tuple[LedgerEntry | None, str | None]:
    """행(셀 7개) → LedgerEntry. 실패 시 (None, 사유)."""
    if len(cells) < 7:
        return None, f"{rownum}행: 컬럼 7개 필요({_BULK_COLS}) — 현재 {len(cells)}개"
    date_c, market_c, name_c, code_c, side_c, qty_c, price_c = cells[:7]
    date = _cell_date(date_c)
    if not _valid_ymd(date):
        return None, f"{rownum}행: 거래일 오류 '{date_c}' (YYYYMMDD, 실제 날짜)"
    market = str(market_c or "").strip() or prices.DEFAULT_MARKET
    name = str(name_c or "").strip()
    code = _cell_code(code_c, market)
    if not (code or name):
        return None, f"{rownum}행: 종목명/코드 둘 다 비어 있음"
    sd = "매도" if str(side_c or "").strip().lower() in ("매도", "sell", "s", "1") else "매수"
    try:
        qty = int(float(str(qty_c).replace(",", "")))
        price = float(str(price_c).replace(",", ""))
    except (ValueError, TypeError):
        return None, f"{rownum}행: 수량/단가 숫자 오류 ('{qty_c}', '{price_c}')"
    if qty <= 0 or price < 0:
        return None, f"{rownum}행: 수량은 1 이상, 단가는 0 이상"
    tag = str(cells[7]).strip() if len(cells) > 7 else ""
    return LedgerEntry(
        date=date, stk_cd=code or name, stk_nm=name or code,
        side=sd, qty=qty, price=price, market=market, tag=tag,
    ), None


def _is_header_row(cells: list[Any]) -> bool:
    first = str(cells[0]).strip() if cells else ""
    return first in ("거래일", "date", "Date", "DATE")


def parse_bulk_entries(text: str) -> tuple[list[LedgerEntry], list[str]]:
    """여러 줄 텍스트(엑셀 복사 등) → LedgerEntry 목록.

    한 줄 = 한 항목, 컬럼 순서 `거래일 시장 종목명 코드 구분 수량 단가`,
    탭 또는 콤마 구분. 잘못된 줄은 건너뛰고 사유를 errors 로 반환.
    """
    out: list[LedgerEntry] = []
    errors: list[str] = []
    for i, raw in enumerate(text.splitlines(), 1):
        s = raw.strip()
        if not s:
            continue
        cells = [p.strip() for p in (s.split("\t") if "\t" in s else s.split(","))]
        if _is_header_row(cells):
            continue
        entry, err = _entry_from_cells(cells, i)
        if entry is not None:
            out.append(entry)
        elif err:
            errors.append(err)
    return out, errors


def parse_excel_entries(path: str) -> tuple[list[LedgerEntry], list[str]]:
    """.xlsx/.csv 파일 → LedgerEntry 목록. 헤더행은 자동 무시."""
    p = Path(path)
    if p.suffix.lower() == ".csv":
        import csv
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    else:
        try:
            import openpyxl
        except ImportError as e:
            raise RuntimeError("엑셀(.xlsx) 처리에는 openpyxl 이 필요합니다.") from e
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

    out: list[LedgerEntry] = []
    errors: list[str] = []
    for i, row in enumerate(rows, 1):
        cells = list(row)
        if all(str(c or "").strip() == "" for c in cells):
            continue
        if _is_header_row(cells):
            continue
        entry, err = _entry_from_cells(cells, i)
        if entry is not None:
            out.append(entry)
        elif err:
            errors.append(err)
    return out, errors


def write_sample_xlsx(path: str) -> Path:
    """원장 업로드용 샘플 .xlsx 생성 (헤더 + 예시 행). 코드열은 텍스트 서식."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원장"
    ws.append(BULK_HEADER)
    sample = [
        ["20260115", "KOSPI", "삼성전자", "005930", "매수", 10, 70000, "장투"],
        ["20260116", "KOSPI", "삼성전자", "005930", "매도", 4, 80000, "장투"],
        ["20260116", "KOSDAQ", "에코프로비엠", "247540", "매수", 5, 120000, "테마"],
        ["20260116", "NASDAQ", "Apple", "AAPL", "매수", 3, 200, "해외"],
    ]
    for r in sample:
        ws.append(r)
    for col, width in zip("ABCDEFGH", (12, 10, 14, 10, 6, 8, 10, 8)):
        ws.column_dimensions[col].width = width
    for cell in ws["D"]:                # 코드열 텍스트 서식(앞자리 0 보존)
        cell.number_format = "@"
    out = Path(path)
    wb.save(out)
    return out


# ── replay ───────────────────────────────────────────────────────────────────
@dataclass
class _Pos:
    qty: int = 0
    cost: float = 0.0          # 잔여 취득원가 (qty * 평단)
    name: str = ""
    market: str = prices.DEFAULT_MARKET


@dataclass
class _DayAgg:
    buy_amt: float = 0.0
    sell_amt: float = 0.0
    realized: float = 0.0
    sold_cost: float = 0.0     # 매도분의 취득원가 (수익률 계산용)


def _replay(entries: list[LedgerEntry], upto_ymd: str) -> tuple[dict[str, _Pos], dict[str, _DayAgg]]:
    """upto_ymd 까지 원장을 재생. 종목별 잔여 포지션과 'upto_ymd 당일' 집계를 반환."""
    through = sorted((e for e in entries if e.date and e.date <= upto_ymd), key=lambda e: e.date)
    pos: dict[str, _Pos] = {}
    day: dict[str, _DayAgg] = {}
    for e in through:
        code = e.stk_cd or e.stk_nm
        p = pos.setdefault(code, _Pos())
        p.name = e.stk_nm or p.name
        p.market = e.market or p.market
        is_today = e.date == upto_ymd
        d = day.setdefault(code, _DayAgg()) if is_today else _DayAgg()
        if e.side == "매수":
            p.qty += e.qty
            p.cost += e.qty * e.price
            d.buy_amt += e.qty * e.price
        else:  # 매도 (보유 초과분은 보유수량으로 클램프)
            avg = (p.cost / p.qty) if p.qty > 0 else e.price
            sell_qty = min(e.qty, p.qty) if p.qty > 0 else 0
            realized = sell_qty * (e.price - avg)
            p.qty -= sell_qty
            p.cost -= sell_qty * avg
            d.sell_amt += e.qty * e.price
            d.realized += realized
            d.sold_cost += sell_qty * avg
    return pos, day


def held_symbols(entries: list[LedgerEntry], upto_ymd: str) -> list[tuple[str, str]]:
    """upto_ymd 기준 보유(qty>0) 종목의 (code, market) 목록 — 시세 조회 대상."""
    pos, _ = _replay(entries, upto_ymd)
    return [(code, p.market) for code, p in pos.items() if p.qty > 0]


# ── 빌드 ─────────────────────────────────────────────────────────────────────
def build_manual_daily(
    date_ymd: str,
    entries: list[LedgerEntry],
    cur_prices: dict[str, float] | None = None,
    fx: dict[str, float] | None = None,
    account_no: str = "수동 원장",
) -> DailyJournalData:
    """원장 + 현재가 → DailyJournalData (기존 렌더 파이프라인과 호환)."""
    cur_prices = cur_prices or {}
    pos, day = _replay(entries, date_ymd)

    # 체결 내역(당일)
    trades: list[Trade] = []
    for e in sorted((x for x in entries if x.date == date_ymd), key=lambda x: x.stk_nm):
        trades.append(Trade(
            ord_no="", stk_cd=e.stk_cd, stk_nm=e.stk_nm, side=e.side,
            ord_qty=e.qty, ord_price=int(round(e.price)),
            cntr_qty=e.qty, cntr_price=int(round(e.price)),
            cntr_time="", venue=e.market, currency=e.currency,
        ))

    # 종목별 실현손익(당일)
    stock_pls: list[StockPL] = []
    for code, d in day.items():
        if d.buy_amt == 0 and d.sell_amt == 0:
            continue
        p = pos.get(code, _Pos())
        rr = (d.realized / d.sold_cost * 100.0) if d.sold_cost > 0 else 0.0
        stock_pls.append(StockPL(
            stk_cd=code, stk_nm=p.name,
            buy_amt=int(round(d.buy_amt)), sell_amt=int(round(d.sell_amt)),
            realized_pl=int(round(d.realized)), return_rate=rr,
            currency=prices.currency_of(p.market),
        ))

    # 보유 종목(replay 종료 상태)
    holdings: list[Holding] = []
    for code, p in pos.items():
        if p.qty <= 0:
            continue
        avg = p.cost / p.qty
        cur = cur_prices.get(code)
        cur_price = float(cur) if cur is not None else avg
        eval_amt = p.qty * cur_price
        cost = p.cost
        pl = eval_amt - cost
        holdings.append(Holding(
            stk_cd=code, stk_nm=p.name, qty=p.qty,
            avg_price=int(round(avg)), cur_price=int(round(cur_price)),
            eval_amt=int(round(eval_amt)), pl_amt=int(round(pl)),
            return_rate=(pl / cost * 100.0) if cost > 0 else 0.0,
            currency=prices.currency_of(p.market),
        ))

    total_realized = sum(s.realized_pl for s in stock_pls)
    total_buy = sum(s.buy_amt for s in stock_pls)
    total_sell = sum(s.sell_amt for s in stock_pls)
    total_eval = sum(h.eval_amt for h in holdings)
    total_eval_pl = sum(h.pl_amt for h in holdings)

    return DailyJournalData(
        date=ymd_dashed(date_ymd),
        account_no=account_no,
        trades=trades,
        stock_pls=sorted(stock_pls, key=lambda s: s.realized_pl, reverse=True),
        holdings=sorted(holdings, key=lambda h: h.eval_amt, reverse=True),
        total_realized_pl=total_realized,
        total_buy_amt=total_buy,
        total_sell_amt=total_sell,
        total_eval_amt=total_eval,
        total_eval_pl=total_eval_pl,
        deposit=0,
        raw={"source": "manual", "fx": fx or {}, "prices": cur_prices},
    )


def _fetch_domestic_via_broker(
    cfg: Config, codes: list[str], api: str, log: LogFn
) -> dict[str, float]:
    """국내 현재가를 키움/한투 API 로 조회. 실패 종목은 생략(상위에서 Yahoo/수동 폴백)."""
    broker = Broker.KIWOOM if api == "kiwoom" else Broker.KIS
    creds = cfg.creds.get(broker)
    if not creds or not creds.app_key or not creds.secret_key:
        raise RuntimeError(f"{broker.display_name} APP KEY/SECRET 가 설정되지 않았습니다.")
    host = broker_info(broker).hosts[cfg.env]
    out: dict[str, float] = {}
    client = make_client(
        broker, host=host, app_key=creds.app_key, secret_key=creds.secret_key,
        is_mock=cfg.is_mock, **creds.extra,
    )
    with client:
        for code in codes:
            try:
                p = client.current_price(code)
                if p and p > 0:
                    out[code] = float(p)
            except Exception as e:  # noqa: BLE001  개별 종목 실패는 폴백으로 처리
                log(f"[warn] {code} {broker.display_name} 현재가 실패: {e}")
    return out


# ── 오케스트레이션 ───────────────────────────────────────────────────────────
def run_manual_journal(
    cfg: Config,
    ymd: str,
    fmt: str = "md",
    *,
    do_fetch: bool = True,
    entries: list[LedgerEntry] | None = None,
    ledger: Ledger | None = None,
    out_label: str = "manual",
    account_no: str = "수동 원장",
    log: LogFn | None = None,
) -> JournalResult:
    """수동 원장 매매일지 생성. cfg 는 AI/journal_dir/시세키 용도로 사용.

    ledger 가 주어지면 해당 계좌 원장을 쓰고(없으면 활성 계좌), entries 가
    주어지면 그 부분집합만으로 재생(체크된 항목만 생성용). out_label 은 출력
    파일명 접미사(계좌별 충돌 방지), account_no 는 일지에 표기될 계좌명.
    """
    log = log or (lambda _m: None)
    if fmt not in ("md", "pdf", "both"):
        raise RuntimeError(f"알 수 없는 형식: {fmt}")
    if len(ymd) != 8 or not ymd.isdigit():
        raise RuntimeError(f"날짜 형식 오류: {ymd} (YYYYMMDD 필요)")

    ledger = ledger if ledger is not None else load_ledger()
    use_entries = ledger.entries if entries is None else entries
    if not use_entries:
        raise RuntimeError("대상 항목이 없습니다. [수동 원장] 탭에서 항목을 추가/선택하세요.")

    log(f"[1/4] {ymd_dashed(ymd)} 기준 원장 재생...")
    symbols = held_symbols(use_entries, ymd)
    domestic = [(c, m) for c, m in symbols if prices.currency_of(m) == "KRW"]
    foreign = [(c, m) for c, m in symbols if prices.currency_of(m) != "KRW"]

    cur_prices: dict[str, float] = dict(ledger.prices)
    fx: dict[str, float] = {}
    domestic_api = (cfg.manual_domestic_api or "yahoo").lower()
    if do_fetch and symbols:
        # 해외주식: 항상 Yahoo
        if foreign:
            log(f"[2/4] 해외 {len(foreign)}종목 Yahoo 현재가 조회...")
            fp, _e = prices.resolve_prices(foreign, ledger.prices, do_fetch=True, log=log)
            cur_prices.update(fp)
        # 국내주식: 설정된 API (yahoo/kiwoom/kis)
        if domestic:
            label = {"yahoo": "Yahoo", "kiwoom": "키움", "kis": "한투"}.get(domestic_api, "Yahoo")
            log(f"[2/4] 국내 {len(domestic)}종목 {label} 현재가 조회...")
            dcodes = [c for c, _m in domestic]
            if domestic_api in ("kiwoom", "kis"):
                try:
                    dp = _fetch_domestic_via_broker(cfg, dcodes, domestic_api, log)
                except Exception as e:  # noqa: BLE001  인증/네트워크 실패 → Yahoo 폴백
                    log(f"[warn] {label} 조회 실패({e}) → Yahoo 폴백")
                    dp, _e = prices.resolve_prices(domestic, ledger.prices, do_fetch=True, log=log)
                else:
                    missing = [(c, m) for c, m in domestic if c not in dp]
                    if missing:  # 일부 실패분은 Yahoo 로 보완
                        yp, _e = prices.resolve_prices(missing, ledger.prices, do_fetch=True, log=log)
                        dp.update(yp)
                cur_prices.update(dp)
            else:
                dp, _e = prices.resolve_prices(domestic, ledger.prices, do_fetch=True, log=log)
                cur_prices.update(dp)
        currencies = sorted({prices.currency_of(m) for _c, m in symbols})
        fx = prices.resolve_fx(currencies, do_fetch=True, log=log)
    else:
        log("[2/4] 시세 조회 생략 (수동값/원가 사용)")

    data = build_manual_daily(ymd, use_entries, cur_prices, fx, account_no=account_no)
    log(f"  → 보유 {len(data.holdings)}종목, 당일 체결 {data.trade_count}건, "
        f"실현손익 {data.total_realized_pl:,}")

    log("[3/4] AI 코멘트 처리...")
    commentary = _maybe_ai(cfg, data, log)

    log(f"[4/4] 출력 생성 ({fmt})...")
    out_base = cfg.journal_dir / f"{ymd}_{out_label}"
    out_json = cfg.journal_dir / f"{ymd}_{out_label}.json"
    cfg.journal_dir.mkdir(parents=True, exist_ok=True)
    snapshot = {
        "ymd": ymd,
        "entries": [e.to_dict() for e in use_entries],
        "resolved_prices": cur_prices,
        "fx": fx,
    }
    out_json.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path, pdf_path = _write_outputs(out_base, data, commentary, fmt)

    log("완료.")
    return JournalResult(
        json_path=out_json, md_path=md_path, pdf_path=pdf_path,
        data=data, commentary=commentary,
    )


__all__ = [
    "LedgerEntry", "Ledger", "LedgerBook", "DEFAULT_ACCOUNT",
    "load_ledger", "save_ledger", "load_book", "save_book",
    "export_book", "import_book",
    "parse_bulk_entries", "parse_excel_entries", "write_sample_xlsx", "BULK_HEADER",
    "build_manual_daily", "held_symbols", "run_manual_journal", "today_ymd",
]
